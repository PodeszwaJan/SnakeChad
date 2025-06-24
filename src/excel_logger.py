import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# --- MODIFIED: Define path to the logs directory ---
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGS_DIR = os.path.join(PROJECT_ROOT, 'logs')

class ExcelLogger:
    def __init__(self, filename_prefix="training_log", sheet_name="Training Data"):
        # --- MODIFIED: Ensure the logs directory exists ---
        if not os.path.exists(LOGS_DIR):
            os.makedirs(LOGS_DIR)
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"{filename_prefix}_{timestamp}.xlsx"
        self.filename = os.path.join(LOGS_DIR, log_filename) # <-- MODIFIED: Save in logs/
        
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = sheet_name
        
        self.headers = ["Episode", "Score", "Best Score", "Epsilon", "Total Reward", "Steps", "Model Type", "Model Name"]
        self._write_header()

    def _write_header(self):
        self.sheet.append(self.headers)
        for i, header_text in enumerate(self.headers, 1):
            cell = self.sheet.cell(row=1, column=i)
            cell.font = Font(bold=True)
            self.sheet.column_dimensions[get_column_letter(i)].width = len(header_text) + 5

    def log_episode(self, data):
        row_data = [data.get(header, "") for header in self.headers]
        self.sheet.append(row_data)

    def save(self):
        self.workbook.save(self.filename)
        print(f"\nTraining data saved to: {self.filename}")