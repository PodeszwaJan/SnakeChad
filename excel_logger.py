import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class ExcelLogger:
    """
    A simple logger to save training data to an Excel (.xlsx) file.
    """
    def __init__(self, filename_prefix="training_log", sheet_name="Training Data"):
        """
        Initializes the logger. Creates a new workbook and sets up the header row.
        A timestamp is added to the filename to prevent overwriting.
        """
        # Create a unique filename with a timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = sheet_name
        
        # Define and write headers
        self.headers = [
            "Episode", "Score", "Best Score", "Epsilon", "Total Reward", "Steps"
        ]
        self._write_header()

    def _write_header(self):
        """Writes the header row to the sheet and formats it."""
        self.sheet.append(self.headers)
        # Make headers bold and auto-fit column width
        for i, header_text in enumerate(self.headers, 1):
            cell = self.sheet.cell(row=1, column=i)
            cell.font = Font(bold=True)
            # Adjust column width for better readability
            self.sheet.column_dimensions[get_column_letter(i)].width = len(header_text) + 5

    def log_episode(self, data):
        """
        Logs a new row of data for an episode.
        
        Args:
            data (dict): A dictionary where keys match the headers.
        """
        # Ensure data is in the correct order based on headers
        row_data = [data.get(header, "") for header in self.headers]
        self.sheet.append(row_data)

    def save(self):
        """Saves the workbook to the specified file."""
        self.workbook.save(self.filename)
        print(f"\nTraining data saved to: {self.filename}")